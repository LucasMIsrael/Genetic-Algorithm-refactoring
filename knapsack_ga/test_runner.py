import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from utils import gerar_instancia
from genetic_algorithm import algoritmo_genetico
import time
import statistics

def testar_algoritmo_genetico(tamanhos, num_geracoes=100, num_execucoes=10):
    resultados = []

    for tam in tamanhos:
        valores_totais = []
        pesos_totais = []
        tempos = []

        for _ in range(num_execucoes):
            pesos, valores, capacidade = gerar_instancia(tam)
            inicio = time.time()
            resultado = algoritmo_genetico(pesos, valores, capacidade, tam_pop=50, num_geracoes=num_geracoes)
            fim = time.time()

            valores_totais.append(resultado["valor_total"])
            pesos_totais.append(resultado["peso_total"])
            tempos.append(round(fim - inicio, 4))

        resultados.append({
            "num_itens": tam,
            "media_valor": round(statistics.mean(valores_totais), 2),
            "desvio_valor": round(statistics.stdev(valores_totais), 2) if len(valores_totais) > 1 else 0,
            "melhor_valor": max(valores_totais),
            "pior_valor": min(valores_totais),
            "media_peso": round(statistics.mean(pesos_totais), 2),
            "media_tempo": round(statistics.mean(tempos), 4),
            "geracoes": num_geracoes,
            "execucoes": num_execucoes
        })
    return resultados

def salvar_resultados_excel(resultados, filename="results/benchmarks.xlsx"):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Resultados"

    fieldnames = [
        "Número de Itens", "Média Valor", "Desvio Valor", "Melhor Valor", "Pior Valor",
        "Média Peso", "Tempo Médio (s)", "Gerações", "Execuções"
    ]

    for col_num, header in enumerate(fieldnames, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num, r in enumerate(resultados, 2):
        valores = [
            r["num_itens"], r["media_valor"], r["desvio_valor"],
            r["melhor_valor"], r["pior_valor"], r["media_peso"],
            r["media_tempo"], r["geracoes"], r["execucoes"]
        ]
        for col_num, val in enumerate(valores, 1):
            sheet.cell(row=row_num, column=col_num, value=val)

    for col_num in range(1, len(fieldnames) + 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 20

    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in sheet.iter_rows(min_row=1, max_row=len(resultados) + 1, min_col=1, max_col=len(fieldnames)):
        for cell in row:
            cell.border = border

    wb.save(filename)

if __name__ == "__main__":
    tamanhos_teste = [5, 10, 1000, 10000]
    resultados = testar_algoritmo_genetico(tamanhos_teste, num_geracoes=100, num_execucoes=10)
    for r in resultados:
        print(r)
    salvar_resultados_excel(resultados)